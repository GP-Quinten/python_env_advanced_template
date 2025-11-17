import json
import logging
import click
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from p03_evaluate_extractions import costs_estimation

logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def load_metrics_json(file_path):
    """Load metrics from a JSON file."""
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def format_percentage(value):
    """Format a float as a percentage with 1 decimal place."""
    # Return the raw float value instead of formatted string
    return value


def format_percentage_with_2_decimals(value):
    """Format a float as a percentage with 2 decimal places for Details column."""
    return f"{value*100:.2f}%"


def format_euros(value):
    """Format a float as euros with 2 decimal places."""
    # Return the raw float value instead of formatted string
    return value


def format_registry_related_breakdown(data):
    """Format detailed breakdown for registry_related."""
    bm = data["binary_metrics"]
    lines = []
    lines.append(f"precision: {format_percentage(bm['precision'])}")
    lines.append(f"recall: {format_percentage(bm['recall'])}")
    lines.append(f"f1_score: {format_percentage(bm['f1_score'])}")
    lines.append(f"accuracy: {format_percentage(bm['accuracy'])}")
    return "\n".join(lines)


def format_regular_field_breakdown(data):
    """Format detailed breakdown for regular fields."""
    summary = data["summary"]
    rb = data["reason_breakdown"]
    correct = rb["correct_extractions"]
    incorrect = rb["incorrect_extractions"]
    total = summary["total_samples"]
    lines = []
    lines.append(
        f"Correct extractions: {correct['total']} ({format_percentage_with_2_decimals(correct['percentage'])})"
    )
    for reason, count in correct["reasons"].items():
        pct = count / total if total > 0 else 0
        lines.append(f"  {reason}: {count} ({format_percentage_with_2_decimals(pct)})")
    lines.append(
        f"Incorrect extractions: {incorrect['total']} ({format_percentage_with_2_decimals(incorrect['percentage'])})"
    )
    for reason, count in incorrect["reasons"].items():
        pct = count / total if total > 0 else 0
        lines.append(f"  {reason}: {count} ({format_percentage_with_2_decimals(pct)})")
    return "\n".join(lines)


def get_accuracy(data):
    """Extract accuracy value from the data."""
    if "summary" in data:
        return data["summary"]["accuracy"]
    else:
        return data["binary_metrics"]["accuracy"]


def format_correct_extractions_ratio(data):
    """Format correct extractions ratio (correct/total)."""
    if "summary" in data:
        correct = data["summary"]["correct_extractions"]
        total = data["summary"]["total_samples"]
        return f"{correct}/{total}"
    else:
        # For registry_related metrics: true_positive + true_negative / total
        correct = (
            data["binary_metrics"]["true_positive"]
            + data["binary_metrics"]["true_negative"]
        )
        total = data["total_samples"]
        return f"{correct}/{total}"


def estimate_cost_from_records(records_jsonl_path, model_name):
    """Estimate costs from a JSONL records file using costs_estimation module."""
    # Map the model name to the costs_estimation model name
    if "small" in model_name.lower():
        model = "mistral-small-latest"
    elif "medium" in model_name.lower():
        model = "mistral-medium-latest"
    else:
        model = "mistral-medium-latest"  # Default to medium if unknown

    # Load the records
    prompts_list = []
    outputs_list = []
    with open(records_jsonl_path, "r", encoding="utf-8") as f:
        for line in f:
            record = json.loads(line)
            prompts_list.append(record["prompt"])

            # Handle different response formats from different models
            if "choices" in record["llm_response"]:
                content = record["llm_response"]["choices"][0]["message"]["content"]
                outputs_list.append(content)
            elif (
                isinstance(record["llm_response"], dict)
                and "content" in record["llm_response"]
            ):
                outputs_list.append(record["llm_response"]["content"])
            else:
                # Default for other formats
                outputs_list.append(str(record["llm_response"]))

    # Estimate costs
    result = costs_estimation.estimate(
        prompts=prompts_list, generations=outputs_list, model=model
    )

    # Return the total cost and the number of records
    return {
        "total_cost": result["total_cost"],
        "num_records": len(prompts_list),
        "cost_details": result,
    }


def style_excel(worksheet):
    """Apply styling to the Excel worksheet."""
    header_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Style all cells with borders
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Style headers (rows 1-2)
    for row in range(1, 3):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    # Style model headers
    for col in range(4, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    # Style the metrics and costs
    for row in range(3, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            if col == 1:  # First column (fields)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(bold=True)
            elif col == 2 or col == 3:  # Info and note columns
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
            else:  # Data columns
                if "%" in str(cell.value) or (
                    "cost" in str(worksheet.cell(row=2, column=col).value).lower()
                    and cell.value is not None
                ):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif "Details" in str(worksheet.cell(row=2, column=col).value):
                    cell.alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths
    worksheet.column_dimensions["A"].width = 18  # Field
    worksheet.column_dimensions["B"].width = 15  # Info
    worksheet.column_dimensions["C"].width = 25  # Note

    # Set column widths for model data
    for i, col in enumerate(range(4, worksheet.max_column + 1)):
        col_letter = get_column_letter(col)
        col_type = (
            i % 5
        )  # 5 columns per model: % kept, total cost, recall, precision, f1-score
        if col_type == 2 or col_type == 4:  # Details or Costs
            worksheet.column_dimensions[col_letter].width = 30
        else:
            worksheet.column_dimensions[col_letter].width = 15


@click.command()
@click.option(
    "--perf_metrics_jsons",
    type=click.Path(exists=True),
    required=True,
    multiple=True,
    help="Paths to performance metrics JSON files for individual fields.",
)
@click.option(
    "--llm_records_jsonls",
    type=click.Path(exists=True),
    required=True,
    multiple=True,
    help="Paths to LLM records JSONL files for individual fields cost estimation.",
)
@click.option(
    "--registry_related_perf_jsons",
    type=click.Path(exists=True),
    required=True,
    multiple=True,
    help="Paths to registry_related performance metrics JSON files.",
)
@click.option(
    "--llm_records_jsonls_registry_related",
    type=click.Path(exists=True),
    required=True,
    multiple=True,
    help="Paths to LLM records JSONL files for registry_related cost estimation.",
)
@click.option(
    "--output_excel",
    type=str,
    required=True,
    help="Path to save the Excel output file.",
)
def main(
    perf_metrics_jsons,
    llm_records_jsonls,
    registry_related_perf_jsons,
    llm_records_jsonls_registry_related,
    output_excel,
):
    """
    Aggregate performance metrics and costs from all models and fields into a formatted Excel report.
    Unlike W02, this workflow processes each field individually with a field-specific prompt,
    potentially leading to better performance and different costs per field.
    """
    start_time = pd.Timestamp.now()
    logger.warning(f"Starting aggregation at {start_time}")

    # Ensure output directory exists
    out_dir = Path(output_excel).parent
    out_dir.mkdir(parents=True, exist_ok=True)

    # Define the models and fields we're working with
    models = ["small_mistral", "medium_mistral"]
    regular_fields = [
        "registry_name",
        "geographical_area",
        "medical_condition",
        "outcome_measure",
    ]

    # Load all performance metrics
    perf_metrics_data = {}
    for file in perf_metrics_jsons:
        data = load_metrics_json(file)
        field = data["field"]
        model = data["model"]
        perf_metrics_data[(field, model)] = data
        logger.warning(f"Loaded performance metrics for {field} / {model}")

    # Load registry-related performance metrics
    registry_related_perf_data = {}
    for file in registry_related_perf_jsons:
        data = load_metrics_json(file)
        model_name = data["model"]
        if "small" in model_name.lower():
            model = "small_mistral"
        elif "medium" in model_name.lower():
            model = "medium_mistral"
        else:
            model = "unknown_model"
        registry_related_perf_data[model] = data
        logger.warning(f"Loaded registry-related performance metrics for {model}")

    # Calculate costs for regular fields - now each field has its own cost
    field_costs = {}
    for file in llm_records_jsonls:
        path = Path(file)
        # Folder structure is .../field/model/field_records.jsonl
        field = path.parent.parent.name
        model = path.parent.name
        cost_data = estimate_cost_from_records(file, model)
        field_costs[(field, model)] = cost_data
        logger.warning(
            f"Calculated costs for {model} / {field}: {cost_data['total_cost']:.4f} euros ({cost_data['num_records']} records)"
        )

    # Calculate costs for registry-related
    registry_related_costs = {}
    for file in llm_records_jsonls_registry_related:
        path = Path(file)
        model = path.parent.name
        cost_data = estimate_cost_from_records(file, model)
        registry_related_costs[model] = cost_data
        logger.warning(
            f"Calculated costs for {model} registry-related: {cost_data['total_cost']:.4f} euros ({cost_data['num_records']} records)"
        )

    # Build the output table as a list of lists, matching the template structure
    rows = []
    # Header row (empty, empty, empty, SMALL mistral x6, empty, medium mistral x6)
    header = [
        "",
        "",
        "",
        "SMALL mistral",
        "",
        "",
        "",
        "",
        "",
        "MEDIUM mistral",
        "",
        "",
        "",
        "",
        "",
    ]
    rows.append(header)

    # Second row with column headers: (empty, empty, Registry related..., % kept, total cost, etc.)
    reg_label = "Registry related\n(73% publications are registry related)"
    header_row = [
        "1",
        "field-specific prompts",
        reg_label,
        "% kept",
        "total cost",
        "recall",
        "precision",
        "f1-score",
        "",
        "% kept",
        "total cost",
        "recall",
        "precision",
        "f1-score",
    ]
    rows.append(header_row)

    # Third row with actual values for registry related for both models
    reg_values_row = ["", "", ""]
    # SMALL mistral and medium mistral metrics
    for model in ["small_mistral", "medium_mistral"]:
        if model in registry_related_perf_data:
            reg_data = registry_related_perf_data[model]
            total_samples = reg_data.get("total_samples", 0)
            true_pos = reg_data["binary_metrics"]["true_positive"]
            false_pos = reg_data["binary_metrics"]["false_positive"]
            predicted_pos_pct = (
                (true_pos + false_pos) / total_samples if total_samples > 0 else 0
            )
            cost_data = registry_related_costs.get(model, None)
            cost_per_record = (
                cost_data["total_cost"] / cost_data["num_records"]
                if cost_data and cost_data["num_records"] > 0
                else 0
            )
            total_cost_220k = cost_per_record * 220000
            reg_values_row.extend(
                [
                    format_percentage(predicted_pos_pct),
                    format_euros(total_cost_220k),
                    format_percentage(reg_data["binary_metrics"]["recall"]),
                    format_percentage(reg_data["binary_metrics"]["precision"]),
                    format_percentage(reg_data["binary_metrics"]["f1_score"]),
                ]
            )
        else:
            reg_values_row.extend(["", "", "", "", ""])
        # Add empty column after each model block except the last
        if model == "small_mistral":
            reg_values_row.append("")
    rows.append(reg_values_row)

    # Fourth row: all empty (visual separation)
    rows.append([""] * 15)

    # Fifth row: section header for other fields
    other_fields_row = [
        "",
        "",
        "Other fields:",
        "cost on 220K",
        "cost for publications left",
        "accuracy",
        "counts correct answers",
        "Details",
        "",
        "cost on 220K",
        "cost for publications left",
        "accuracy",
        "counts correct answers",
        "Details",
    ]
    rows.append(other_fields_row)

    # Rows for each field
    total_costs = {"small_mistral": 0, "medium_mistral": 0}
    for field in [
        "registry_name",
        "geographical_area",
        "medical_condition",
        "outcome_measure",
    ]:
        row = ["", "", field]
        for model in ["small_mistral", "medium_mistral"]:
            metrics_key = (field, model)
            if metrics_key in perf_metrics_data:
                metrics = perf_metrics_data[metrics_key]
                accuracy = get_accuracy(metrics)
                counts = format_correct_extractions_ratio(metrics)
                details = format_regular_field_breakdown(metrics)

                # Get field-specific costs
                cost_data = field_costs.get(metrics_key, None)
                cost_per_record = (
                    cost_data["total_cost"] / cost_data["num_records"]
                    if cost_data and cost_data["num_records"] > 0
                    else 0
                )
                total_cost_220k = cost_per_record * 220000
                # Get the registry_related percentage calculated earlier for this model
                registry_related_pct = 0
                if model in registry_related_perf_data:
                    reg_data = registry_related_perf_data[model]
                    total_samples = reg_data.get("total_samples", 0)
                    true_pos = reg_data["binary_metrics"]["true_positive"]
                    false_pos = reg_data["binary_metrics"]["false_positive"]
                    registry_related_pct = (
                        (true_pos + false_pos) / total_samples
                        if total_samples > 0
                        else 0
                    )
                cost_for_reg_related = cost_per_record * 220000 * registry_related_pct
                total_costs[model] += cost_for_reg_related

                row.extend(
                    [
                        format_euros(total_cost_220k),
                        format_euros(cost_for_reg_related),
                        format_percentage(accuracy),
                        counts,
                        details,
                    ]
                )
            else:
                row.extend(["", "", "", "", ""])
            if model == "small_mistral":
                row.append("")
        rows.append(row)

    # Final row: TOTAL COST
    total_row = ["", "", ""]
    for model in ["small_mistral", "medium_mistral"]:
        # Add registry-related cost to total
        if model in registry_related_costs:
            cost_data = registry_related_costs[model]
            cost_per_record = (
                cost_data["total_cost"] / cost_data["num_records"]
                if cost_data["num_records"] > 0
                else 0
            )
            registry_cost_220k = cost_per_record * 220000
            total_costs[model] += registry_cost_220k
        total_row.append("TOTAL COST:")
        total_row.extend([format_euros(total_costs[model]), "", "", "", ""])
        if model == "small_mistral":
            total_row.append("")
    rows.append(total_row)

    # Set columns explicitly to match template
    columns = [
        "",
        "Unnamed: 1",
        "Unnamed: 2",
        "sm_% kept",
        "sm_total cost",
        "sm_recall",
        "sm_precision",
        "sm_f1-score",
        "",
        "lg_% kept",
        "lg_total cost",
        "lg_recall",
        "lg_precision",
        "lg_f1-score",
        "",
    ]

    # Create DataFrame
    df = pd.DataFrame(rows)

    # Write to Excel
    with pd.ExcelWriter(output_excel, engine="openpyxl") as wb:
        df.to_excel(wb, index=False, header=False)
        worksheet = wb.sheets["Sheet1"]

        # --- Advanced Formatting ---
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        # 1. Merge A2:A9 and B2:B9
        worksheet.merge_cells("A2:A9")
        worksheet["A2"].value = "<N°>"
        worksheet.merge_cells("B2:B9")
        worksheet["B2"].value = "<prompt description>"

        # 2. Merge C2:C3 and format
        worksheet.merge_cells("C2:C3")
        worksheet["C2"].value = (
            "**Registry related**\n  (73% publications are registry related)"
        )
        # Format '(73% publications are registry related)' in font size 8 italic
        from openpyxl.styles import Font

        cell = worksheet["C2"]
        cell.font = Font(size=11, bold=True)

        # 3. Merge D1:H1 and J1:N1
        worksheet.merge_cells("D1:H1")
        worksheet["D1"].value = "SMALL mistral"
        worksheet.merge_cells("J1:N1")
        worksheet["J1"].value = "MEDIUM mistral"

        # 4. Wrap text everywhere, set column widths wide enough
        for col in range(1, 15 + 1):
            col_letter = get_column_letter(col)
            for row in range(1, worksheet.max_row + 1):
                worksheet[f"{col_letter}{row}"].alignment = Alignment(
                    wrap_text=True, vertical="center"
                )
            worksheet.column_dimensions[col_letter].width = 22

        # 5. Increase indent for C5:C8
        for row in range(5, 9):
            worksheet[f"C{row}"].alignment = Alignment(
                indent=2, horizontal="left", vertical="center", wrap_text=True
            )

        # 6. Alignment
        # D9, J9: align right, bold
        worksheet["D9"].alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=True
        )
        worksheet["D9"].font = Font(bold=True)
        worksheet["J9"].alignment = Alignment(
            horizontal="right", vertical="center", wrap_text=True
        )
        worksheet["J9"].font = Font(bold=True)
        # C2:C8: align left
        for row in range(2, 9):
            worksheet[f"C{row}"].alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
        # E9, K9: align left
        worksheet["E9"].alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        worksheet["K9"].alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        # All other cells: center
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.alignment is None or (
                    cell.alignment.horizontal not in ["left", "right"]
                ):
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

        # 7. Format 'Details' cells
        def format_details_cell(cell):
            if cell.value and "Correct extractions:" in str(cell.value):
                lines = cell.value.split("\n")
                formatted = []
                for line in lines:
                    if line.startswith("Correct extractions:") or line.startswith(
                        "Incorrect extractions:"
                    ):
                        formatted.append(line)
                    else:
                        formatted.append("      " + line)
                cell.value = "\n".join(formatted)

        for row in range(4, worksheet.max_row + 1):
            for col in [8, 14]:  # Details columns
                cell = worksheet.cell(row=row, column=col)
                format_details_cell(cell)

        # Make all header cells bold
        worksheet["A2"].font = Font(bold=True)
        worksheet["B2"].font = Font(bold=True)

        # C2 to C8
        for row in range(2, 9):
            worksheet[f"C{row}"].font = Font(bold=True)

        # D1, J1 (model headers)
        worksheet["D1"].font = Font(bold=True)
        worksheet["J1"].font = Font(bold=True)

        # D2 to H2, J2 to N2 (column headers)
        for col in range(4, 9):
            worksheet.cell(row=2, column=col).font = Font(bold=True)
        for col in range(10, 15):
            worksheet.cell(row=2, column=col).font = Font(bold=True)

    logger.warning(f"Saved performance metrics and costs to {output_excel}")
    end_time = pd.Timestamp.now()
    elapsed = (end_time - start_time).total_seconds()
    logger.warning(f"Completed aggregation in {elapsed:.1f} seconds")


if __name__ == "__main__":
    main()
