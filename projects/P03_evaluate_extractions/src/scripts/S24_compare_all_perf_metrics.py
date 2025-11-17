import json
import logging
import click
from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.WARNING,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)


def load_metrics_json(file_path):
    """Load metrics from a JSON file."""
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def format_percentage(value):
    """Format a float as a percentage with 1 decimal place."""
    return f"{value*100:.1f}%"


def format_regular_field_breakdown(data):
    """Format detailed breakdown for regular fields."""
    summary = data["summary"]
    rb = data["reason_breakdown"]
    correct = rb["correct_extractions"]
    incorrect = rb["incorrect_extractions"]
    total = summary["total_samples"]
    lines = []
    lines.append(
        f"Correct extractions: {correct['total']} ({format_percentage(correct['percentage'])})"
    )
    for reason, count in correct["reasons"].items():
        pct = count / total if total > 0 else 0
        lines.append(f"  {reason}: {count} ({format_percentage(pct)})")
    lines.append(
        f"Incorrect extractions: {incorrect['total']} ({format_percentage(incorrect['percentage'])})"
    )
    for reason, count in incorrect["reasons"].items():
        pct = count / total if total > 0 else 0
        lines.append(f"  {reason}: {count} ({format_percentage(pct)})")
    return "\n".join(lines)


def format_registry_related_breakdown(data):
    """Format detailed breakdown for registry_related."""
    bm = data["binary_metrics"]
    lines = []
    lines.append(f"precision: {format_percentage(bm['precision'])}")
    lines.append(f"recall: {format_percentage(bm['recall'])}")
    lines.append(f"f1_score: {format_percentage(bm['f1_score'])}")
    return "\n".join(lines)


def format_correct_extractions_ratio(data):
    """Format correct extractions ratio (correct/total)."""
    if "summary" in data:
        correct = data["summary"]["correct_extractions"]
        total = data["summary"]["total_samples"]
        return f"{correct}/{total}"
    else:
        # For registry_related metrics: true_positive + true_negative / total
        correct = (
            data["binary_metrics"]["true_positive"]
            + data["binary_metrics"]["true_negative"]
        )
        total = data["total_samples"]
        return f"{correct}/{total}"


def get_accuracy(data):
    """Extract accuracy value from the data."""
    if "summary" in data:
        return data["summary"]["accuracy"]
    else:
        return data["binary_metrics"]["accuracy"]


def style_excel(workbook):
    """Apply styling to the Excel worksheet."""
    ws = workbook.active
    header_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    max_col = ws.max_column

    # Style header rows (first two rows)
    for row in range(1, 3):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    # Style data rows
    for row in range(3, ws.max_row + 1):
        # Field column
        cell = ws.cell(row=row, column=1)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # Total Samples column
        cell = ws.cell(row=row, column=2)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style other columns
        for col in range(3, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border

            # Column types repeat in groups of 3
            col_type = (col - 3) % 3

            if col_type == 0:  # Accuracy columns
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_type == 1:  # Counts columns
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:  # Detail columns
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

    # Set column widths
    ws.column_dimensions["A"].width = 20  # Field
    ws.column_dimensions["B"].width = 15  # Total Samples

    # Set column widths for model data (groups of 3 columns)
    for col in range(3, max_col + 1):
        col_type = (col - 3) % 3
        if col_type == 0 or col_type == 1:  # Accuracy or Counts
            ws.column_dimensions[get_column_letter(col)].width = 15
        else:  # Details
            ws.column_dimensions[get_column_letter(col)].width = 40

    return workbook


@click.command()
@click.option(
    "--perf_metrics_jsons",
    type=click.Path(exists=True),
    required=True,
    multiple=True,
    help="Paths to performance metrics JSON files.",
)
@click.option(
    "--registry_related_perf_jsons",
    type=click.Path(exists=True),
    required=True,
    multiple=True,
    help="Paths to registry_related performance metrics JSON files.",
)
@click.option(
    "--output_excel",
    type=str,
    required=True,
    help="Path to save the Excel output file.",
)
def main(perf_metrics_jsons, registry_related_perf_jsons, output_excel):
    """
    Aggregate performance metrics from all models and fields into a formatted Excel report.
    Processes 12 regular metrics JSONs and 3 registry_related JSONs.
    """
    start_time = pd.Timestamp.now()
    logger.warning(f"Starting aggregation at {start_time}")

    models = ["small_mistral", "large_mistral", "gpt4o_openai"]

    # Define fields in correct order to ensure they appear in the same order in the Excel file
    regular_fields = [
        "medical_condition",
        "outcome_measure",
        "geographical_area",
        "registry_name",
    ]
    all_fields = regular_fields + ["registry_related"]

    # Create a structured dictionary to store data by field and model
    data_by_field_model = {}

    # Initialize empty data structure for all fields and models
    for field in all_fields:
        data_by_field_model[field] = {}
        for model in models:
            data_by_field_model[field][model] = None

    # Load regular field metrics (12 files)
    for file in perf_metrics_jsons:
        data = load_metrics_json(file)
        field = data["field"]
        model = data["model"]
        data_by_field_model[field][model] = data
        logger.warning(
            f"Loaded regular metrics for field '{field}' and model '{model}'"
        )

    # Load registry_related metrics (3 files)
    for file in registry_related_perf_jsons:
        data = load_metrics_json(file)
        # Extract model from file path if not in data
        # Registry metrics often have a different model name format
        model_name = data.get("model", "")

        # Map model name to standard names if needed
        if "small" in model_name.lower():
            model = "small_mistral"
        elif "large" in model_name.lower():
            model = "large_mistral"
        elif "gpt4o" in model_name.lower() or "openai" in model_name.lower():
            model = "gpt4o_openai"
        else:
            # Extract from file path as fallback
            path = Path(file)
            parts = path.parts
            if "small_mistral" in parts:
                model = "small_mistral"
            elif "large_mistral" in parts:
                model = "large_mistral"
            elif "gpt4o_openai" in parts:
                model = "gpt4o_openai"
            else:
                # If we can't determine model, use filename
                model = path.parent.name

        field = "registry_related"
        data_by_field_model[field][model] = data
        logger.warning(f"Loaded registry_related metrics for model '{model}'")

    # Debug data loading before creating rows
    logger.warning(f"Fields in data_by_field_model: {list(data_by_field_model.keys())}")
    for field in data_by_field_model:
        logger.warning(
            f"Models for field {field}: {list(data_by_field_model[field].keys())}"
        )

    # Create Excel file directly with openpyxl
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Performance Metrics"

    # Create header structure manually
    # First row: empty cells for A1, B1, then model names merged across 3 columns each
    for i, m in enumerate(models):
        col_start = 3 + i * 3
        worksheet.merge_cells(
            start_row=1, start_column=col_start, end_row=1, end_column=col_start + 2
        )
        worksheet.cell(row=1, column=col_start).value = m

    # Second row: Field, Total Samples, then Accuracy, Counts, Details for each model
    worksheet.cell(row=2, column=1).value = "Field"
    worksheet.cell(row=2, column=2).value = "Total Samples"

    for i in range(len(models)):
        base_col = 3 + i * 3
        worksheet.cell(row=2, column=base_col).value = "Accuracy"
        worksheet.cell(row=2, column=base_col + 1).value = "Counts correct answers"
        worksheet.cell(row=2, column=base_col + 2).value = "Details"

    # Add data rows
    row_idx = 3  # Start from row 3
    for field in all_fields:
        logger.warning(f"Processing field: {field}")

        # Field name in column A
        worksheet.cell(row=row_idx, column=1).value = field

        # Get total samples from first available model with data
        total_samples = "N/A"
        for m in models:
            if (
                m in data_by_field_model[field]
                and data_by_field_model[field][m] is not None
            ):
                if field == "registry_related":
                    total_samples = data_by_field_model[field][m].get(
                        "total_samples", "N/A"
                    )
                else:
                    total_samples = data_by_field_model[field][m]["summary"][
                        "total_samples"
                    ]
                break

        # Total samples in column B
        worksheet.cell(row=row_idx, column=2).value = total_samples

        # For each model, add Accuracy, Ratio and Detail breakdown
        for i, m in enumerate(models):
            base_col = 3 + i * 3

            if (
                m in data_by_field_model[field]
                and data_by_field_model[field][m] is not None
            ):
                d = data_by_field_model[field][m]
                acc = get_accuracy(d)

                # Accuracy column
                worksheet.cell(row=row_idx, column=base_col).value = format_percentage(
                    acc
                )

                # Counts correct answers column
                worksheet.cell(row=row_idx, column=base_col + 1).value = (
                    format_correct_extractions_ratio(d)
                )

                # Details column
                if field == "registry_related":
                    worksheet.cell(row=row_idx, column=base_col + 2).value = (
                        format_registry_related_breakdown(d)
                    )
                else:
                    worksheet.cell(row=row_idx, column=base_col + 2).value = (
                        format_regular_field_breakdown(d)
                    )
            else:
                worksheet.cell(row=row_idx, column=base_col).value = "N/A"
                worksheet.cell(row=row_idx, column=base_col + 1).value = "N/A"
                worksheet.cell(row=row_idx, column=base_col + 2).value = "N/A"

        # Move to next row
        row_idx += 1
        # logger.warning(f"Added row {row_idx-1}: {field}")

    # Apply styling
    style_excel(workbook)

    # Ensure output directory exists
    out_dir = Path(output_excel).parent
    out_dir.mkdir(parents=True, exist_ok=True)

    # Save the workbook
    workbook.save(output_excel)

    logger.warning(f"Saved aggregate metrics to {output_excel}")
    logger.warning(
        f"Excel file contains {len(all_fields)} rows with fields in order: {', '.join(all_fields)}"
    )
    end_time = pd.Timestamp.now()
    elapsed = (end_time - start_time).total_seconds()
    logger.warning(f"Completed aggregation in {elapsed:.1f} seconds")


if __name__ == "__main__":
    main()
